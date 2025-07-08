#!/usr/bin/env python3
"""
Create a comprehensive sample Excel file with instructions
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_sample_with_instructions():
    """Create a sample Excel file with detailed instructions"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Create Instructions sheet
    instructions_ws = wb.active
    instructions_ws.title = "Instructions"
    
    # Add instructions content
    instructions_content = [
        ["SEO Alt Text Generator - Instructions", "", "", ""],
        ["", "", "", ""],
        ["How to use this tool:", "", "", ""],
        ["1. Your Excel file must have a column with image URLs", "", "", ""],
        ["2. Column names that work for image URLs:", "", "", ""],
        ["   - image_url", "", "", ""],
        ["   - image", "", "", ""],
        ["   - url", "", "", ""],
        ["   - image_link", "", "", ""],
        ["", "", "", ""],
        ["3. Optional: Alt text column (will be created if missing)", "", "", ""],
        ["   Column names that work for alt text:", "", "", ""],
        ["   - alt_text", "", "", ""],
        ["   - alt", "", "", ""],
        ["   - description", "", "", ""],
        ["   - alt_description", "", "", ""],
        ["", "", "", ""],
        ["4. You can include any other columns with additional data", "", "", ""],
        ["", "", "", ""],
        ["5. The tool will:", "", "", ""],
        ["   - Automatically detect your image URL column", "", "", ""],
        ["   - Create an alt_text column if one doesn't exist", "", "", ""],
        ["   - Generate SEO-friendly alt text for missing entries", "", "", ""],
        ["   - Allow you to edit any generated text", "", "", ""],
        ["   - Export the updated file for download", "", "", ""],
        ["", "", "", ""],
        ["See the 'Sample Data' sheet for an example format", "", "", ""],
    ]
    
    # Add instructions to sheet
    for row_idx, row_data in enumerate(instructions_content, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = instructions_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Title row
                cell.font = Font(bold=True, size=16)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, size=16, color="FFFFFF")
            elif value.startswith(("1.", "2.", "3.", "4.", "5.")):  # Main points
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            elif value.startswith("   -"):  # Sub-points
                cell.font = Font(italic=True)
                cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    
    # Create Sample Data sheet
    sample_ws = wb.create_sheet("Sample Data")
    
    # Create headers
    headers = ['ID', 'Product Name', 'image_url', 'alt_text', 'Category', 'Price']
    for col, header in enumerate(headers, 1):
        cell = sample_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Sample data
    sample_data = [
        [1, "Red Sports Car", "https://images.unsplash.com/photo-1494976688153-018c804d0dd7?w=800", "", "Automotive", "$45,000"],
        [2, "Mountain Landscape", "https://images.unsplash.com/photo-1506905925346-21bda4d32df4?w=800", "Beautiful mountain scenery with snow-capped peaks", "Nature", "N/A"],
        [3, "Coffee Cup", "https://images.unsplash.com/photo-1541167760496-1628856ab772?w=800", "", "Food & Drink", "$12.99"],
        [4, "Modern Office", "https://images.unsplash.com/photo-1497366216548-37526070297c?w=800", "", "Architecture", "N/A"],
        [5, "Golden Retriever", "https://images.unsplash.com/photo-1552053831-71594a27632d?w=800", "", "Animals", "N/A"],
        [6, "Fresh Vegetables", "https://images.unsplash.com/photo-1540420773420-3366772f4999?w=800", "Colorful fresh vegetables including tomatoes, peppers, and leafy greens", "Food", "$8.50"],
        [7, "City Skyline", "https://images.unsplash.com/photo-1449824913935-59a10b8d2000?w=800", "", "Urban", "N/A"],
        [8, "Yoga Class", "https://images.unsplash.com/photo-1506629905117-b5d1b8b6c4e5?w=800", "", "Health & Wellness", "$25.00"],
    ]
    
    # Add sample data to worksheet
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = sample_ws.cell(row=row_idx, column=col_idx, value=value)
            if not value:  # Empty alt text cells
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif col_idx == 4 and value:  # Existing alt text
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
    
    # Auto-adjust column widths for both sheets
    for ws in [instructions_ws, sample_ws]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save the file
    filename = "sample_images.xlsx"
    wb.save(filename)
    print(f"Enhanced sample Excel file created: {filename}")
    
    # Print statistics
    total_rows = len(sample_data)
    missing_alt_text = sum(1 for row in sample_data if not row[3])
    print(f"Total rows: {total_rows}")
    print(f"Rows missing alt text: {missing_alt_text}")
    print(f"Rows with existing alt text: {total_rows - missing_alt_text}")

if __name__ == "__main__":
    create_sample_with_instructions()