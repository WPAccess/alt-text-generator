#!/usr/bin/env python3
"""
Create a sample Excel file with working image URLs
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

def create_working_sample():
    """Create a sample Excel file with working image URLs"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Create Instructions sheet
    instructions_ws = wb.active
    instructions_ws.title = "Instructions"
    
    # Add instructions content
    instructions_content = [
        ["SEO Alt Text Generator - Instructions", "", "", ""],
        ["", "", "", ""],
        ["How to use this tool:", "", "", ""],
        ["1. Your Excel file must have a column with image URLs", "", "", ""],
        ["2. Column names that work for image URLs:", "", "", ""],
        ["   - image_url", "", "", ""],
        ["   - image", "", "", ""],
        ["   - url", "", "", ""],
        ["   - image_link", "", "", ""],
        ["", "", "", ""],
        ["3. Optional: Alt text column (will be created if missing)", "", "", ""],
        ["   Column names that work for alt text:", "", "", ""],
        ["   - alt_text", "", "", ""],
        ["   - alt", "", "", ""],
        ["   - description", "", "", ""],
        ["   - alt_description", "", "", ""],
        ["", "", "", ""],
        ["4. You can include any other columns with additional data", "", "", ""],
        ["", "", "", ""],
        ["IMPORTANT: Use direct image URLs, not webpage URLs", "", "", ""],
        ["✓ Good: https://picsum.photos/800/600", "", "", ""],
        ["✓ Good: https://via.placeholder.com/800x600", "", "", ""],
        ["✗ Bad: https://unsplash.com/photos/photo-name-xxx", "", "", ""],
        ["", "", "", ""],
        ["5. The tool will:", "", "", ""],
        ["   - Automatically detect your image URL column", "", "", ""],
        ["   - Create an alt_text column if one doesn't exist", "", "", ""],
        ["   - Generate SEO-friendly alt text for missing entries", "", "", ""],
        ["   - Allow you to edit any generated text", "", "", ""],
        ["   - Export the updated file for download", "", "", ""],
        ["", "", "", ""],
        ["See the 'Sample Data' sheet for an example format", "", "", ""],
    ]
    
    # Add instructions to sheet
    for row_idx, row_data in enumerate(instructions_content, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = instructions_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Title row
                cell.font = Font(bold=True, size=16)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, size=16, color="FFFFFF")
            elif value.startswith(("1.", "2.", "3.", "4.", "5.")):  # Main points
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            elif value.startswith("   -"):  # Sub-points
                cell.font = Font(italic=True)
                cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
            elif value.startswith("✓"):  # Good examples
                cell.font = Font(color="006600")
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            elif value.startswith("✗"):  # Bad examples
                cell.font = Font(color="CC0000")
                cell.fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")
    
    # Create Sample Data sheet
    sample_ws = wb.create_sheet("Sample Data")
    
    # Create headers
    headers = ['ID', 'Product Name', 'image_url', 'alt_text', 'Category', 'Price']
    for col, header in enumerate(headers, 1):
        cell = sample_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Sample data with working image URLs
    sample_data = [
        [1, "Random Nature Image", "https://picsum.photos/800/600?random=1", "", "Nature", "$25.00"],
        [2, "Technology Product", "https://picsum.photos/800/600?random=2", "Modern technology device on clean background", "Technology", "$199.99"],
        [3, "Abstract Art", "https://picsum.photos/800/600?random=3", "", "Art", "$89.50"],
        [4, "Architecture Photo", "https://picsum.photos/800/600?random=4", "", "Architecture", "N/A"],
        [5, "Landscape View", "https://picsum.photos/800/600?random=5", "", "Nature", "N/A"],
        [6, "Urban Scene", "https://picsum.photos/800/600?random=6", "Busy city street with modern buildings", "Urban", "$15.00"],
        [7, "Portrait Photography", "https://picsum.photos/800/600?random=7", "", "Portrait", "$45.00"],
        [8, "Food Photography", "https://picsum.photos/800/600?random=8", "", "Food", "$12.99"],
    ]
    
    # Add sample data to worksheet
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = sample_ws.cell(row=row_idx, column=col_idx, value=value)
            if not value:  # Empty alt text cells
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif col_idx == 4 and value:  # Existing alt text
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
    
    # Auto-adjust column widths for both sheets
    for ws in [instructions_ws, sample_ws]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save the file
    filename = "sample_images.xlsx"
    wb.save(filename)
    print(f"Working sample Excel file created: {filename}")
    
    # Print statistics
    total_rows = len(sample_data)
    missing_alt_text = sum(1 for row in sample_data if not row[3])
    print(f"Total rows: {total_rows}")
    print(f"Rows missing alt text: {missing_alt_text}")
    print(f"Rows with existing alt text: {total_rows - missing_alt_text}")

if __name__ == "__main__":
    create_working_sample()