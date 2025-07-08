#!/usr/bin/env python3
"""
Script to create a sample Excel file for testing the Alt Text Generator
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

def create_sample_excel():
    """Create a sample Excel file with image URLs for testing"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Image Data"
    
    # Create headers
    headers = ['ID', 'Product Name', 'image_url', 'alt_text', 'Category', 'Price']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
    
    # Add instruction row
    instruction_row = [
        "Sample ID", 
        "Your product name", 
        "REQUIRED: Put your image URLs here", 
        "OPTIONAL: Leave empty to generate with AI", 
        "Any additional data", 
        "Any additional data"
    ]
    for col, instruction in enumerate(instruction_row, 1):
        cell = ws.cell(row=2, column=col, value=instruction)
        cell.font = Font(italic=True, color="666666")
        cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    
    # Sample data with real image URLs from various sources
    sample_data = [
        [1, "Red Sports Car", "https://images.unsplash.com/photo-1494976688153-018c804d0dd7?w=800", "", "Automotive", "$45,000"],
        [2, "Mountain Landscape", "https://images.unsplash.com/photo-1506905925346-21bda4d32df4?w=800", "Beautiful mountain scenery with snow-capped peaks", "Nature", "N/A"],
        [3, "Coffee Cup", "https://images.unsplash.com/photo-1541167760496-1628856ab772?w=800", "", "Food & Drink", "$12.99"],
        [4, "Modern Office", "https://images.unsplash.com/photo-1497366216548-37526070297c?w=800", "", "Architecture", "N/A"],
        [5, "Golden Retriever", "https://images.unsplash.com/photo-1552053831-71594a27632d?w=800", "", "Animals", "N/A"],
        [6, "Fresh Vegetables", "https://images.unsplash.com/photo-1540420773420-3366772f4999?w=800", "Colorful fresh vegetables including tomatoes, peppers, and leafy greens", "Food", "$8.50"],
        [7, "City Skyline", "https://images.unsplash.com/photo-1449824913935-59a10b8d2000?w=800", "", "Urban", "N/A"],
        [8, "Yoga Class", "https://images.unsplash.com/photo-1506629905117-b5d1b8b6c4e5?w=800", "", "Health & Wellness", "$25.00"],
        [9, "Beach Sunset", "https://images.unsplash.com/photo-1507525428034-b723cf961d3e?w=800", "", "Nature", "N/A"],
        [10, "Laptop Computer", "https://images.unsplash.com/photo-1496181133206-80ce9b88a853?w=800", "", "Technology", "$1,299.99"]
    ]
    
    # Add sample data to worksheet (starting from row 3 now)
    for row_idx, row_data in enumerate(sample_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the file
    filename = "sample_images.xlsx"
    wb.save(filename)
    print(f"Sample Excel file created: {filename}")
    
    # Print statistics
    total_rows = len(sample_data)
    missing_alt_text = sum(1 for row in sample_data if not row[3])  # Alt text is at index 3
    print(f"Total rows: {total_rows}")
    print(f"Rows missing alt text: {missing_alt_text}")
    print(f"Rows with existing alt text: {total_rows - missing_alt_text}")

if __name__ == "__main__":
    create_sample_excel()