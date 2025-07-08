#!/usr/bin/env python3
"""
Create a sample Excel file with working image URLs
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

def create_final_sample():
    """Create a sample Excel file with working image URLs"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Create Instructions sheet
    instructions_ws = wb.active
    instructions_ws.title = "Instructions"
    
    # Add instructions content
    instructions_content = [
        ["SEO Alt Text Generator - Instructions", "", "", ""],
        ["", "", "", ""],
        ["How to use this tool:", "", "", ""],
        ["1. Your Excel file must have a column with image URLs", "", "", ""],
        ["2. Column names that work for image URLs:", "", "", ""],
        ["   - image_url", "", "", ""],
        ["   - image", "", "", ""],
        ["   - url", "", "", ""],
        ["   - image_link", "", "", ""],
        ["", "", "", ""],
        ["3. Optional: Alt text column (will be created if missing)", "", "", ""],
        ["   Column names that work for alt text:", "", "", ""],
        ["   - alt_text", "", "", ""],
        ["   - alt", "", "", ""],
        ["   - description", "", "", ""],
        ["   - alt_description", "", "", ""],
        ["", "", "", ""],
        ["4. You can include any other columns with additional data", "", "", ""],
        ["", "", "", ""],
        ["IMPORTANT: Use direct image URLs, not webpage URLs", "", "", ""],
        ["✓ Good: Direct links to image files (.jpg, .png, .webp)", "", "", ""],
        ["✗ Bad: Links to web pages containing images", "", "", ""],
        ["", "", "", ""],
        ["5. The tool will:", "", "", ""],
        ["   - Automatically detect your image URL column", "", "", ""],
        ["   - Create an alt_text column if one doesn't exist", "", "", ""],
        ["   - Generate SEO-friendly alt text for missing entries", "", "", ""],
        ["   - Allow you to edit any generated text", "", "", ""],
        ["   - Export the updated file for download", "", "", ""],
        ["", "", "", ""],
        ["See the 'Sample Data' sheet for an example format", "", "", ""],
        ["Replace the sample URLs with your actual image URLs", "", "", ""],
    ]
    
    # Add instructions to sheet
    for row_idx, row_data in enumerate(instructions_content, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = instructions_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Title row
                cell.font = Font(bold=True, size=16)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, size=16, color="FFFFFF")
            elif value.startswith(("1.", "2.", "3.", "4.", "5.")):  # Main points
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            elif value.startswith("   -"):  # Sub-points
                cell.font = Font(italic=True)
                cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
            elif value.startswith("✓"):  # Good examples
                cell.font = Font(color="006600")
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            elif value.startswith("✗"):  # Bad examples
                cell.font = Font(color="CC0000")
                cell.fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")
    
    # Create Sample Data sheet
    sample_ws = wb.create_sheet("Sample Data")
    
    # Create headers
    headers = ['ID', 'Product Name', 'image_url', 'alt_text', 'Category', 'Price']
    for col, header in enumerate(headers, 1):
        cell = sample_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
    
    # Sample data with placeholder URLs (users will replace with their own)
    sample_data = [
        [1, "Product Photo 1", "https://example.com/image1.jpg", "", "Products", "$25.00"],
        [2, "Lifestyle Image", "https://example.com/image2.jpg", "People enjoying outdoor activities", "Lifestyle", "$199.99"],
        [3, "Brand Logo", "https://example.com/logo.png", "", "Branding", "$89.50"],
        [4, "Building Photo", "https://example.com/building.jpg", "", "Architecture", "N/A"],
        [5, "Nature Scene", "https://example.com/nature.jpg", "", "Nature", "N/A"],
        [6, "Tech Product", "https://example.com/tech.jpg", "Modern smartphone with sleek design", "Technology", "$599.00"],
        [7, "Food Image", "https://example.com/food.jpg", "", "Food", "$12.99"],
        [8, "Team Photo", "https://example.com/team.jpg", "", "Corporate", "N/A"],
    ]
    
    # Add sample data to worksheet
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = sample_ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 3:  # Image URL column
                cell.fill = PatternFill(start_color="E8F3FF", end_color="E8F3FF", fill_type="solid")
                cell.font = Font(color="0066CC")
            elif not value and col_idx == 4:  # Empty alt text cells
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif col_idx == 4 and value:  # Existing alt text
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
    
    # Add note about replacing URLs
    note_row = len(sample_data) + 3
    sample_ws.cell(row=note_row, column=1, value="NOTE:")
    sample_ws.cell(row=note_row, column=2, value="Replace the example.com URLs above with your actual image URLs")
    sample_ws.cell(row=note_row, column=1).font = Font(bold=True, color="CC0000")
    sample_ws.cell(row=note_row, column=2).font = Font(italic=True, color="CC0000")
    
    # Auto-adjust column widths for both sheets
    for ws in [instructions_ws, sample_ws]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save the file
    filename = "sample_images.xlsx"
    wb.save(filename)
    print(f"Sample Excel file created: {filename}")
    
    # Print statistics
    total_rows = len(sample_data)
    missing_alt_text = sum(1 for row in sample_data if not row[3])
    print(f"Total rows: {total_rows}")
    print(f"Rows missing alt text: {missing_alt_text}")
    print(f"Rows with existing alt text: {total_rows - missing_alt_text}")
    print("Note: Replace example.com URLs with your actual image URLs")

if __name__ == "__main__":
    create_final_sample()