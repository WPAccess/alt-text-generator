import openpyxl
from openpyxl.styles import Font, PatternFill
import logging

logger = logging.getLogger(__name__)

class ExcelProcessor:
    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = None
        self.worksheet = None
        self.header_row = None
        self.image_url_column = None
        self.alt_text_column = None
        
    def read_excel(self):
        """Read Excel file and return data structure"""
        try:
            self.workbook = openpyxl.load_workbook(self.filepath)
            self.worksheet = self.workbook.active
            
            # Find header row and identify columns
            self._identify_columns()
            
            if not self.image_url_column:
                raise ValueError("Could not find image URL column. Expected column names: 'image_url', 'image', 'url', 'image_link'")
            
            data = []
            for row_num in range(self.header_row + 1, self.worksheet.max_row + 1):
                row_data = self._get_row_data(row_num)
                if row_data:
                    data.append(row_data)
            
            return data
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            raise
    
    def _identify_columns(self):
        """Identify header row and relevant columns"""
        # Look for header row in first 10 rows
        for row_num in range(1, min(11, self.worksheet.max_row + 1)):
            row = self.worksheet[row_num]
            headers = [cell.value.lower() if cell.value else "" for cell in row]
            
            # Look for image URL column
            image_url_indicators = ['image_url', 'image', 'url', 'image_link', 'imageurl']
            for col_num, header in enumerate(headers, 1):
                if any(indicator in header for indicator in image_url_indicators):
                    self.image_url_column = col_num
                    self.header_row = row_num
                    break
            
            if self.image_url_column:
                # Look for alt text column
                alt_text_indicators = ['alt_text', 'alt', 'description', 'alt_description', 'alttext']
                for col_num, header in enumerate(headers, 1):
                    if any(indicator in header for indicator in alt_text_indicators):
                        self.alt_text_column = col_num
                        break
                
                # If no alt text column found, create one
                if not self.alt_text_column:
                    self.alt_text_column = len(headers) + 1
                    self.worksheet.cell(row=self.header_row, column=self.alt_text_column, value="alt_text")
                
                break
    
    def _get_row_data(self, row_num):
        """Get data for a specific row"""
        try:
            image_url = self.worksheet.cell(row=row_num, column=self.image_url_column).value
            if not image_url:
                return None
            
            alt_text = self.worksheet.cell(row=row_num, column=self.alt_text_column).value
            
            # Get all other columns for context
            row_data = {}
            for col_num in range(1, self.worksheet.max_column + 1):
                header_cell = self.worksheet.cell(row=self.header_row, column=col_num)
                header = header_cell.value if header_cell.value else f"Column_{col_num}"
                
                cell_value = self.worksheet.cell(row=row_num, column=col_num).value
                row_data[header] = cell_value
            
            return {
                'row_index': row_num,
                'image_url': image_url,
                'alt_text': alt_text,
                'needs_alt_text': not alt_text or str(alt_text).strip() == "",
                'row_data': row_data
            }
            
        except Exception as e:
            logger.error(f"Error getting row data for row {row_num}: {str(e)}")
            return None
    
    def get_image_url(self, row_index):
        """Get image URL for a specific row"""
        try:
            if not self.worksheet or not self.image_url_column:
                logger.error(f"Worksheet or image_url_column not initialized")
                return None
            return self.worksheet.cell(row=row_index, column=self.image_url_column).value
        except Exception as e:
            logger.error(f"Error getting image URL for row {row_index}: {str(e)}")
            return None
    
    def update_excel(self, updated_data, output_path):
        """Update Excel file with new alt text and save"""
        try:
            # Update cells with new alt text
            for row_data in updated_data:
                row_index = row_data['row_index']
                alt_text = row_data.get('alt_text', '')
                
                # Update alt text cell
                cell = self.worksheet.cell(row=row_index, column=self.alt_text_column)
                cell.value = alt_text
                
                # Highlight updated cells
                if alt_text and row_data.get('was_generated', False):
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    cell.font = Font(color="006600")
            
            # Save to output path
            self.workbook.save(output_path)
            logger.info(f"Updated Excel file saved to {output_path}")
            
        except Exception as e:
            logger.error(f"Error updating Excel file: {str(e)}")
            raise
